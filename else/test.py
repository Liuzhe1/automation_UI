import openpyxl
from web_key.keys import keys


def arguments(value):
    data = dict()
    if value:
        str_temp = value.split(';')
        for temp in str_temp:
            t = temp.split('=', 1)
            data[t[0]] = t[1]
    else:
        pass
    return data


excel = openpyxl.load_workbook('D:/test1.xlsx')
# sheet = excel['Sheet1']

for name in excel.sheetnames:
    sheet = excel[name]
    print('******正在执行{}Sheet页******'.format(name))

for values in sheet.values :
    if type(values[0]) is int :
        print('***正在执行：{}***'.format(values[3]))
        data = arguments(values[2])
        print(data)

        if values[1] == 'open_browser':
            key = keys(**data)

        elif 'assert' in values[1]:
            status = getattr(key, values[1])(expected=values[4], **data)
            if status:
                sheet.cell(row=values[0]+1, column=6).value = 'pass'
            else:
                sheet.cell(row=values[0]+1, column=6).value = 'failed'
            excel.save('D:/test1.xlsx')

        else:
            getattr(key,values[1])(**data)

