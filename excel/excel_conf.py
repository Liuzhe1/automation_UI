from openpyxl.styles import PatternFill, Font


def pass_(cell, row, column):
    cell(row=row, column=column).value = 'Pass'
    cell(row=row, column=column).fill = PatternFill('solid', fgColor='AACF91')
    cell(row=row, column=column).font = Font(bold=True)


def fail_(cell, row, column):
    cell(row=row, column=column).value = 'Fail'
    cell(row=row, column=column).fill = PatternFill('solid', fgColor='FF0000')
    cell(row=row, column=column).font = Font(bold=True)
